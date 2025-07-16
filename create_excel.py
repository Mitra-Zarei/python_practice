# INSTRUKTIONER 
# Skapa via Python ett excel-dokument som du populerar med data från CSV-filen sales_data.csv, rubrikerna skall ha fetstil och vara i storlek 14.
# Lägg till en ny kolumn i din excel-fil som visar upp totala försäljningen (dvs. antal * pris per enhet).
# Skapa ett nytt ark där du listar försäljning per region, denna skall vara formatterad som en tabell.
# Lägg till ett passande diagram som visar totala försäljningen per region.


import csv
import openpyxl
import os

# Denna variabel kan du använda för att läsa in filen sales_data.csv som liger i samma mapp som denna fil.
input_path = os.path.join(os.path.dirname(__file__), 'sales_data.csv')
# Denna variabel kan du använda för att skriva ut filen sales_output_v2.xlsx i samma mapp som denna fil.
output_path = os.path.join(os.path.dirname(__file__), 'sales_output_v2.xlsx')

# Defines the names of the worksheets
SALES_SHEET_NAME = 'Sales'
REGION_SHEET_NAME = 'Region'

# Create an Excell_file and an active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = SALES_SHEET_NAME

# Open the CSV file and read the data
with open(input_path) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    data = list(csv_reader)

# Put the headers on the first line 
headers = data[0]
for column_index, header in enumerate(headers , start=1):
    cell = worksheet.cell( row=1 , column=column_index, value=header)

# Write the data for the rest of the document
rows_excluding_heading = data[1:]
for row_index, row_data in enumerate(rows_excluding_heading, start=2):
    for column_index, cell_data in enumerate(row_data, start=1):
        worksheet.cell(row=row_index, column=column_index, value=cell_data)    

# Add column with total sales
worksheet.cell(row=1, column=6, value='Total sales')
for row_index in range(2, worksheet.max_row + 1):
    count = int(worksheet.cell(row=row_index, column=4).value)
    price = float(worksheet.cell(row=row_index, column=5).value)
    total_amount = count * price
    worksheet.cell(row=row_index, column=6, value=total_amount) 

# style in bold and size 14
for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, size=14)


    """ REGION SALES SHEET """
    
# Save sales by region in a dict
region_sales = {}
for row_index in range(2, worksheet.max_row + 1):
    region = worksheet.cell(row=row_index, column=3).value
    total_sales = float(worksheet.cell(row=row_index, column=6).value)
    region_sales[region] = round(region_sales.get(region, 0) + total_sales, 2)
    
region_sheet = workbook.create_sheet(REGION_SHEET_NAME)
worksheet = workbook[region_sheet.title]

# Create the column
worksheet.cell(row=1, column=1, value="Region")
worksheet.cell(row=1, column=2, value="Sales")

# Define how big our sheet should be
table_range = f"A1:B{len(region_sales) + 1}"

# Create the tabell
table = openpyxl.worksheet.table.Table(displayName="Region_Sales", ref=table_range)

# tabell style
style = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium20",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style

# Add the table to the sheet
worksheet.add_table(table)

# Enter the data for sales by region
for index, (name, sales) in enumerate(region_sales.items(), start=2):
    worksheet.cell(row=index, column=1, value=name)
    worksheet.cell(row=index, column=2, value=sales) 

    """ create diagram """

# create a datareferens for sales
data = openpyxl.chart.Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=worksheet.max_row)

# create a categorireferens for regioner
categories = openpyxl.chart.Reference(worksheet, min_col=1, min_row=1, max_row=worksheet.max_row)

# Create a bar graph
chart = openpyxl.chart.BarChart()
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Sale per region"
chart.x_axis.title = "Region"
chart.y_axis.title = "Sale (kr)"

# Add the diagram to the workbook
worksheet.add_chart(chart, "D2")

workbook.save(output_path)

